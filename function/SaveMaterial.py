from openpyxl import Workbook, load_workbook
import os

file_name = "임베디드_재료.xlsx"

if os.path.exists(file_name):
    # 파일이 존재하면 불러오기
    wb = load_workbook(file_name)
    ws = wb.active
else:
    # 파일이 없으면 새로 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "재료"

    column = ['번호', '재료']
    ws.append(column)


# 재료 입력 (공백으로 구분된 여러 재료 입력 가능)
Material = input("재료 목록을 입력하세요 (예: 재료1 재료2 재료3)(입력 후 엔터 클릭): ").split()

# 기존 데이터의 마지막 행부터 순차적으로 재료를 추가
# 기존 행 개수 확인 후, 번호를 이어서 증가시키기 위해 사용
existing_rows = ws.max_row

for idx, mat in enumerate(Material, start=existing_rows):
    ws.append([idx, mat])

# 엑셀 파일 저장
wb.save(file_name)

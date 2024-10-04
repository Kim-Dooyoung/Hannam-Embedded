from openpyxl import Workbook, load_workbook
import os


class MaterialHandler:
    def __init__(self, file_name):
        self.file_name = file_name
        self.materials = []

    def load_or_create(self):
        if os.path.exists(self.file_name):
            # 파일이 존재하면 불러오기
            self.wb = load_workbook(self.file_name)
            self.ws = self.wb.active
        else:
            # 파일이 없으면 새로 생성
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = "재료"
            column = ['번호', '재료']
            self.ws.append(column)

    def add_materials(self, materials):
        # 기존 데이터의 마지막 행부터 순차적으로 재료를 추가
        existing_rows = self.ws.max_row
        for idx, mat in enumerate(materials, start=existing_rows):
            self.ws.append([idx, mat])

    def save(self):
        self.wb.save(self.file_name)

    def get_materials(self):
        return self.materials
